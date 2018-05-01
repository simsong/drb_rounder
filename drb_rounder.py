#!/usr/bin/env python36
#
# Census DRB Rounder
# Implements the rounding rules of the RDCs.
# Includes both code that can be imported into other programs, as well
# as a program that will round input files.

from decimal import Decimal,ROUND_HALF_EVEN,Context
import os
import re
import math

ROUND4_METHOD="round4"
COUNTS_METHOD="counts"

################################################################
### Rounding code
################################################################
def all_spaces(s):
    return all(ch==' ' for ch in s)

def nearest(n, number):
    """Returns n to the nearest number"""
    return math.floor((n / number) + 0.5) * number

LESS_THAN_15 = "<15"
def round_counts(n):
    """Implements the DRB rounding rules for counts. Note that we return a string, 
    so that we can report N < 15"""
    n = int(n)
    assert n == math.floor(n)    # make sure it is an integer; shouldn't be needed with above
    assert n >= 0
    if      0 <= n <      15: return LESS_THAN_15
    if     15 <= n <=     99: return str(nearest( n,   10))
    if    100 <= n <=    999: return str(nearest( n,   50))
    if   1000 <= n <=   9999: return str(nearest( n,  100))
    if  10000 <= n <=  99999: return str(nearest( n,  500))
    if 100000 <= n <= 999999: return str(nearest( n, 1000))
    return round_decimal( n )

prec4_rounder = Context(prec=4, rounding=ROUND_HALF_EVEN)
def round4_float(f):
    return float(str(prec4_rounder.create_decimal(f)))

def round4_str(s,add_spaces=True,round_counts=False):
    """Take a string, convert it to a number, and round it. Add spaces if
    required. If it is not a number, don't round. Puts commas back in the
    numbers. Returns as a string.
    """
    tmp = s

    # Remove the non-digits
    digit_chars = "0123456789.-"
    leading_text = ""
    trailing_text = ""
    while len(tmp)>0 and tmp[0] not in digit_chars:
        leading_text = leading_text + tmp[0]
        tmp = tmp[1:]

    # remove trailing non-digits
    while len(tmp)>0 and tmp[-1] not in digit_chars:
        trailing_text = tmp[-1] + trailing_text
        tmp = tmp[:-1]
    
    # Perform floating point rounding
    # Both of the following are equivalent 
    rounded = round4_float(float(tmp))

    if round_counts and 0 <= rounded <= 7:
        rounded = round_counts(rounded)

    # If the original string had commas, add them back
    # Otherwise just change the rounded number to a string
    if "," in s:
        rounded_str = "{:,}".format(rounded)
    else:
        rounded_str = "{:}".format(rounded)

    # If the original string had no ".", make sure that the new string has no "."
    if ("." not in s) and ("." in rounded_str):
        rounded_str = rounded_str[0: rounded_str.find(".")]
        
    # If the original string ended in ".", make sure that the new string ends in "."
    if (s.endswith(".")) and (not rounded_str.endswith(".")) and ("." in rounded_str):
        rounded_str = rounded_str[0: rounded_str.find(".")+1]

    # Put it back together and return
    ret = leading_text + rounded_str + trailing_text

    if add_spaces:
        # Add missing spaces
        ret += " " * (len(s) - len(ret))
        assert len(s) == len(ret)
    return ret
        
ORIGINAL='original'
ORIGINAL_WITHOUT_COMMAS='original_without_commas'
ORIGINAL_HAS_COMMAS='original_has_commas'
ROUNDING_METHOD='rounding_method'
ROUNDED_WITHOUT_COMMAS='rounded_without_commas'
ROUNDED='rounded'
NEEDS_ROUNDING='needs_rounding'
def analyze_for_rounding(original):
    """Analyze a string for rounding and return a dictionary including original string, rounded string, and rounding method."""
    ret = {}
    ret[ORIGINAL]                = original = str(original) # be sure it is a string
    ret[ORIGINAL_HAS_COMMAS]     = ',' in original
    ret[ORIGINAL_WITHOUT_COMMAS] = original.replace(",","") # remove commas
    if "." in ret[ORIGINAL_WITHOUT_COMMAS]:
        ret[ROUNDING_METHOD] = ROUND4_METHOD
        ret[ROUNDED_WITHOUT_COMMAS] = round4_str(ret[ORIGINAL_WITHOUT_COMMAS])
    else:
        ret[ROUNDING_METHOD] = COUNTS_METHOD
        ret[ROUNDED_WITHOUT_COMMAS] = round_counts(ret[ORIGINAL_WITHOUT_COMMAS])
    # 
    if ret[ORIGINAL_HAS_COMMAS]:
        ret[ROUNDED] = "{:,}".format(float(ret['rounded_without_commas']))
    else:
        ret[ROUNDED] = ret[ROUNDED_WITHOUT_COMMAS]
    ret[NEEDS_ROUNDING]    = ret[ORIGINAL_WITHOUT_COMMAS] != ret[ROUNDED_WITHOUT_COMMAS]
    return ret
    
################################################################
###
### This code is for the drb_rounder application program.
###
################################################################


################################################################
### String manipulation code
################################################################

number_re = re.compile(r"\d{1,3}((,\d\d\d)|(\d\d\d))*([.]\d*)?")
def find_next_number(line,pos=0):
    """Find the next number in line, starting at position pos.
    Returns number starting position and length as (start,end)
    or None."""
    m = number_re.search(line[pos:])
    if m:
        span = m.span()
        return (span[0]+pos,span[1]+pos)

def numbers_in_line(line):
    """Returns an array with all of the numbers in the line"""
    ret = []
    pos = 0
    while True:
        loc = find_next_number(line,pos)
        if not loc:
            break
        ret.append(loc)
        pos = loc[1]
    return ret
    
def safe_open(filename,mode,return_none=False, zap=False):
    """Like open, but if filename exists and mode is 'w', produce an error"""
    if 'w' in mode and os.path.exists(filename) and not zap:
        print("***************************************")
        print("OUTPUT FILE EXISTS: {}".format(filename))
        print("Please delete or rename file and restart program.")
        if return_none:
            return None
        exit(1)
    return open(filename,mode)

class DRBRounder:
    LOGFILE_ROUNDER_EXTENSIONS = ['.log','.txt','.sas']
    HTML_HEADER="""<html>
    <head>
    <style>
    table.no-spacing {
      border-spacing:0; /* Removes the cell spacing via CSS */
      border-collapse: collapse;  /* Optional - if you don't want to have double border where cells touch */
    }

    #line_numbers {
      color: black;
      background-color: LightGray;
      margin-right: 10px;
      text-align: right;
      float:left;
    }
    #content {
      margin-right: 10px;
      float:left;
    }
    .bcount {
      color: DarkRed;
      background-color: yellow;
      font-weight: bold;
    }
    .bfloat {
      color: DarkBlue;
      background-color: yellow;
      font-weight: bold;
    }
    .rcount {
      color: DarkRed;
      background-color: LightGreen;
      font-weight: bold;
    }
    .rfloat {
      color: DarkBlue;
      background-color: LightGreen;
      font-weight: bold;
    }
    </style>
    </head>
    <body>
    <table class='no-spacing'>
    <tr><th class='bcount'> <tt>Count Needing Rounding </tt></th> <th class='rcount'> <tt>Rounded Count </tt></th></tr>
    <tr><th class='bfloat'> <tt>Float Needing Rounding </tt></th> <th class='rfloat'> <tt>Rounded Float </tt></th></tr>
    </table>
    """
    def process_csvfile(self):
        """Process a tab or comma-delimited file and create a rounded file"""
        (basename,ext) = os.path.splitext(self.fname)
        outfilename = basename+"_rounded"+ext
        print("{} --> {}".format(self.fname,outfilename))
        line_number = 0
        with open(self.fname,"rU") as infile:
            with safe_open(outfilename,"w") as outfile:
                for line in infile:
                    stripped_line = line.rstrip()    # remove EOL
                    eol_len = len(line) - len(stripped_line)
                    eol = line[-eol_len:] # preserve the EOL characters

                    line_number += 1
                    delimiter = args.delimiter
                    if delimiter not in stripped_line:
                        print("No delimiter found in line {}: {}".format(line_number,stripped_line))
                        if " " in stripped_line:
                            print("Using space as delimiter for this line")
                            delimiter = " "

                    fields = stripped_line.split(delimiter)
                    rounded_fields = [round_str(field) for field in fields]
                    outfile.write(delimiter.join(rounded_fields))
                    outfile.write(eol) # output the original eol
        print("Lines processed: {}".format(line_number))

    def process_xlsx(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.fname)
        for sheetname in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheetname)
            for cell in sheet.get_cell_collection():
                if type(cell.value)==int:
                    pass
        print(dir(wb))

    def process_logfile(self):
        # Make sure that our intended output files do not exist
        (name,ext) = os.path.splitext(self.fname)
        fname_rounded = name + "_rounded" + ext
        frounded = safe_open(fname_rounded, "w", return_none=True, zap=args.zap)

        # before - highlight items that need rounding
        fname_before  = name + "_0.html"
        fbefore = safe_open(fname_before, "w", return_none=True, zap=args.zap)

        # after - show it rounded
        fname_after  = name + "_1.html"
        fafter = safe_open(fname_after, "w", return_none=True, zap=args.zap)

        if (not frounded) or (not fbefore) or (not fafter):
            print("PROGRAM HALTS")
            exit(1)

        fbefore.write(self.HTML_HEADER)
        fafter.write(self.HTML_HEADER)

        # Note that the output was rounded

        ROUNDING_RULES_MESSAGE = "*** DRB ROUNDING RULES HAVE {}BEEN APPLIED ***"

        frounded.write(ROUNDING_RULES_MESSAGE + "\n")
        fbefore.write("<p><b>" + ROUNDING_RULES_MESSAGE.format("NOT ") + "</b></p>\n")
        fafter.write("<p><b>"  + ROUNDING_RULES_MESSAGE.format("")     + "</b></p>\n")

        # Count the number of lines
        linecount = open(self.fname).read().count("\n")

        # Add the rounding rules
        for what in [fbefore,fafter]:
            what.write("<div id='line_numbers'>\n<pre>\n")
            what.write("".join(["{:}\n".format(line_number) for line_number in range(1,linecount+1)]))
            what.write("</pre>\n</div><div id='content'>\n<pre>\n")
        
        with open(self.fname) as fin:
            for line in fin:
                pos = 0             # where we are on the line

                spans = numbers_in_line(line)
                for span in spans:
                    (col0,col1) = (span[0], span[1]) # the columns where the number appears
                    ret = analyze_for_rounding(line[ col0 : col1 ])
                    if ret[NEEDS_ROUNDING]:
                        if ret[ROUNDING_METHOD]==ROUND4_METHOD:
                            kind='float'
                        else:
                            kind='count'
                            # If we have replaced with LESS_THAN_15, see if we can grab a few more spaces...
                            if ret[ROUNDED] == LESS_THAN_15:
                                extra_spaces = len(LESS_THAN_15) - len(ret[ORIGINAL])
                                if all_spaces( line[col1 : col1 + extra_spaces] ):
                                    col1 = col1 + extra_spaces

                        fbefore.write( line[ pos : col0] ) # part of line before span in question
                        fbefore.write("<span class='b{}'>".format(kind))
                        fbefore.write( line[ col0 : col1] ) #  span in question
                        fbefore.write("</span>")

                        fafter.write( line[ pos : col0] ) # part of line after span in question
                        fafter.write("<span class='r{}'>".format(kind))
                        fafter.write( ret['rounded'] ) #  span in question
                        fafter.write("</span>")

                        frounded.write( line[ pos : col0] )
                        frounded.write( ret['rounded'] )

                        pos = col1 # next character on line to process is col1

                # get end of line (or entire line, if not spans)
                fbefore.write( line[ pos: ] )  
                fafter.write( line[ pos: ] ) 
                frounded.write( line[ pos: ] ) 

        for what in [fbefore,fafter]:
            what.write("</div></body></html>\n")

        frounded.close()
        fbefore.close()
        fafter.close()

    def __init__(self,args,fname):
        self.args  = args
        self.fname = fname

    def process(self):
        if self.args.log:
            self.process_logfile()
            return

        (base,ext) = os.path.splitext(self.fname)
        ext = ext.lower()
        if ext == ".xlsx":
            self.process_xlsx()
        elif ext in ['.csv','.tsv']:
            self.process_csvfile()
        elif ext in self.LOGFILE_ROUNDER_EXTENSIONS:
            self.process_logfile()
        else:
            print("Don't know how to process '{}' type file in: {}".format(ext,fname),out=sys.stderr)
            exit(1)

if __name__=="__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Perform DRB-style rounding',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('files', type=str, nargs='+', help='File[s] to round')
    parser.add_argument("--tab",action="store_true",help="Assume CSV is tab delimited")
    parser.add_argument("--log",help="Invoke logfile rounder. This is the default for files ending in {}".format(" ".join(DRBRounder.LOGFILE_ROUNDER_EXTENSIONS)))
    parser.add_argument("--counts",help="Apply rounding rules for small counts: 0-7 rounds to 4",action='store_true')
    parser.add_argument("--zap",help="overwrite output files",action='store_true')
    args = parser.parse_args()
    args.delimiter = "\t"       # default delimiter
    for fname in args.files:
        d = DRBRounder(args,fname)
        d.process()
