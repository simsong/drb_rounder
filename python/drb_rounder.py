#!/usr/bin/env python36
#
# Census DRB Rounder
# Implements the rounding rules of the RDCs.
# Includes both code that can be imported into other programs, as well
# as a program that will round input files.

# Note: the .xlsx files may not open properly on a Mac due to the fact that Anaconda packages an out-of-date xlsxwriter.
# see: https://xlsxwriter.readthedocs.io/changes.html#changes

import sys
import os

import number
import helpers
import logging
from helpers import LESS_THAN_15, ROUND4_METHOD, COUNTS_METHOD

################################################################
### Logging System
################################################################
def setup_logger(name, log_file, format, log_mode, stream_handler):
    """Function to set up as many loggers as you want"""
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S"

    handler = logging.FileHandler(log_file, mode=log_mode)
    handler.setFormatter(logging.Formatter(fmt=format, datefmt=DATE_FORMAT))

    logger = logging.getLogger(name)
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)
    if stream_handler:
        logger.addHandler(logging.StreamHandler())  # stderr
    return logger

class DRBRounder:
    LOGFILE_ROUNDER_EXTENSIONS = ['.log', '.txt', '.sas', '.lst', '.tex', '.py', '.r']
    HTML_HEADER = """<html>
    <head>
    <style>
    table.no-spacing {
      border-spacing:0; /* Removes the cell spacing via CSS */
      border-collapse: collapse;  /* Optional - if you don't want to have double border where cells touch */
    }

    #line_numbers {
      color: black;
      background-color: LightGray;
      margin-right: 10px;
      text-align: right;
      float:left;
    }
    #content {
      margin-right: 10px;
      float:left;
    }
    .bcount {
      color: DarkRed;
      background-color: yellow;
      font-weight: bold;
    }
    .bfloat {
      color: DarkBlue;
      background-color: yellow;
      font-weight: bold;
    }
    .rcount {
      color: DarkRed;
      background-color: LightGreen;
      font-weight: bold;
    }
    .rfloat {
      color: DarkBlue;
      background-color: LightGreen;
      font-weight: bold;
    }
    </style>
    </head>
    <body>
    <table class='no-spacing'>
    <tr><th class='bcount'> <tt>Count Needing Rounding </tt></th> <th class='rcount'> <tt>Rounded Count </tt></th></tr>
    <tr><th class='bfloat'> <tt>Float Needing Rounding </tt></th> <th class='rfloat'> <tt>Rounded Float </tt></th></tr>
    </table>
    """

    def __init__(self, args, fname):
        self.args  = args
        self.fname = fname
        self.new_fname = os.path.splitext(fname)[0] + "_rounded.xlsx"
        self.logfile_path = os.path.splitext(fname)[0] + "_rounded.log"
        if not self.args.zap:
            error = False
            for fn in [self.new_fname, self.logfile_path]:
                if os.path.exists(fn) and not args.zap:
                    rounder_logger.error("Error: %s exists. Delete file or add --zap option", fn)
                    error = True
            if error:
                exit(1)

        if os.path.exists(self.logfile_path):
            os.unlink(self.logfile_path)
        self.values_logger  = setup_logger('values', self.logfile_path, '%(message)s', 'w', False)

    def process_csvfile(self):
        """Process a tab or comma-delimited file and create a rounded file"""
        # Open file and set up
        (basename, ext) = os.path.splitext(self.fname)
        outfilename = basename + "_rounded" + ext
        line_number = 0
        delimiter = self.args.delimiter

        # Open infile/outfile and begin performing rounding rules
        with open(self.fname, "rU") as infile:
            with helpers.safe_open(outfilename, "w", zap=self.args.zap) as outfile:
                for line in infile:
                    stripped_line = line.rstrip()  # remove EOL
                    eol_len = len(line) - len(stripped_line)
                    if eol_len != 0:  # preserve the EOL characters if needed
                        eol = line[-eol_len:]
                    else:
                        eol = ""

                    line_number += 1

                    # Check to make sure delimiter is seen
                    if delimiter not in stripped_line:
                        rounder_logger.info("\tNo delimiter found in line {}: {}".format(line_number, stripped_line))
                        if " " in stripped_line:
                            rounder_logger.info("\tUsing space as delimiter for this line")
                            delimiter = " "

                    # Split by delimiter and round all items
                    fields = stripped_line.split(delimiter)
                    rounded_fields = []
                    for field in fields:
                        num = number.Number(field)
                        num.round()
                        if num.needed_rounding:
                            self.values_logger.info("line %d  %s --> %s",line_number,num.original,num.rounded)
                            rounded_fields.append(num.rounded)
                        else:
                            rounded_fields.append(num.original)
                    outfile.write(delimiter.join(rounded_fields))
                    outfile.write(eol)  # output the original eol

    def process_xlsx(self, is_xlsx):
        """Process an excel spreadsheet and create a rounded spreadsheet"""
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.comments import Comment

        FILL_ORANGE     = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        FILL_BLUE       = PatternFill(start_color='00b8ff',   end_color='00b8ff', fill_type='solid')
        COMMENT_FLOAT   = Comment('Orange: rounder identified a FLOAT that needed to be rounded', 'DRB_ROUNDER')
        COMMENT_INTEGER = Comment('Blue: rounder identified a COUNT that needed to be rounded', 'DRB_ROUNDER')
        first_float_flag = True
        first_integer_flag = True
        fname = self.fname

        # Self.args will only be None when running tests
        if self.args == None:
            highlight = False
        else:
            highlight = self.args.highlight

        # Make sure the excel spreadsheet exists
        if not os.path.exists(fname):
            rounder_logger.error('ABORT: Excel spreadsheet does not exist')
            sys.exit(1)

        # Convert the workbook if needed
        if is_xlsx == False:
            helpers.convert_to_xlsx(fname)
            (base, ext) = os.path.splitext(fname)
            fname = base + ".xlsx"

        # Make sure no formulas exist in the spreadsheet
        if helpers.check_for_excel_formulas(fname):
            rounder_logger.error('ABORT: Excel spreadsheet contains formulas. '
                                 'Please remove the formulas in the highlighted cells')
            sys.exit(1)

        # Apply rounding rules to all cells in spreadsheet
        wb = load_workbook(fname, data_only=True)  # Open the workbook
        import openpyxl
        openpyxl.writer.excel.save_workbook(wb,self.new_fname)
        abort

        values_seen = 0
        values_rounded = 0
        for sheetname in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheetname)

            # reads each row left to right, then each row, top to bottom
            for cell in sheet.get_cell_collection():  

                values_seen += 1

                # Windows stores excel flaots as single precision sometimes, 
                # which cased a value 0.004237 to be read by Python as 0.0042369999999999994.
                # The following code checks to see if a float is stored and, if it is, it is rounded
                # to 15 significant figures first. IEEE floating point gives imprecision after 17 digits
                if type(cell.value)==float:
                    value = format(cell.value,'.15') 
                    num = number.Number(value, method=ROUND4_METHOD)
                else:
                    num = number.Number(str(cell.value))

                num.round()
                if num.needed_rounding == True:
                    self.values_logger.info("cell %s!%s%s  %s --> %s",sheetname,cell.column,cell.row,num.original,num.rounded)
                    if num.method == ROUND4_METHOD:
                        if not highlight:  # argument passed to only highlight spreadsheet
                            try:  # Successful typecast if no special characters
                                cell.value = float(num.rounded)
                            except ValueError:
                                cell.value = num.rounded
                        cell.fill = FILL_ORANGE
                        if first_float_flag:
                            cell.comment = COMMENT_FLOAT
                            first_float_flag = False
                    elif num.method == COUNTS_METHOD:
                        if not highlight:  # argument passed to only highlight spreadsheet
                            if num.rounded == LESS_THAN_15:
                                cell.value = LESS_THAN_15
                            else:
                                try:  # Successful typecast if no special characters
                                    cell.value = int(num.rounded)
                                except ValueError:
                                    cell.value = num.rounded
                        cell.fill = FILL_BLUE
                        if first_integer_flag:
                            cell.comment = COMMENT_INTEGER
                            first_integer_flag = False

        # Save the rounded calculations to a new file
        try:
            wb.save(self.new_fname)
            rounder_logger.info("STATUS:   New spreadsheet written to %s",self.new_fname)
        except PermissionError:
            rounder_logger.error('ABORT: Could not save. Please close rounded spreadsheet')
            exit(1)

    def process_logfile(self):
        # Make sure that our intended output files do not exist
        (name, ext) = os.path.splitext(self.fname)
        fname_rounded = name + "_rounded" + ext
        frounded = helpers.safe_open(fname_rounded, "w", return_none=True, zap=args.zap)

        # before - highlight items that need rounding
        fname_before = name + "_0.html"
        fbefore = helpers.safe_open(fname_before, "w", return_none=True, zap=args.zap)

        # after - show it rounded
        fname_after = name + "_1.html"
        fafter = helpers.safe_open(fname_after, "w", return_none=True, zap=args.zap)

        if (not frounded) or (not fbefore) or (not fafter):
            rounder_logger.error('ABORT: Could not open log and html files')
            sys.exit(1)

        fbefore.write(self.HTML_HEADER)
        fafter.write(self.HTML_HEADER)

        # Note that the output was rounded
        ROUNDING_RULES_MESSAGE = "*** DRB ROUNDING RULES HAVE {}BEEN APPLIED ***"

        frounded.write(ROUNDING_RULES_MESSAGE + "\n")
        fbefore.write("<p><b>" + ROUNDING_RULES_MESSAGE.format("NOT ") + "</b></p>\n")
        fafter.write("<p><b>"  + ROUNDING_RULES_MESSAGE.format("")     + "</b></p>\n")

        # Count the number of lines
        linecount = open(self.fname).read().count("\n")

        # Add the rounding rules
        for what in [fbefore, fafter]:
            what.write("<div id='line_numbers'>\n<pre>\n")
            what.write("".join(["{:}\n".format(line_number) for line_number in range(1, linecount+1)]))
            what.write("</pre>\n</div><div id='content'>\n<pre>\n")

        with open(self.fname) as fin:
            linenumber = 0
            for line in fin:
                linenumber += 1
                pos = 0  # where we are on the line
                spans = helpers.numbers_in_line(line)
                for span in spans:
                    (col0, col1) = (span[0], span[1])  # the columns where the number appears
                    num = number.Number(line[col0:col1])
                    num.round()
                    if num.needed_rounding:
                        self.values_logger.info("line %s  %s --> %s",linenumber,num.original,num.rounded)
                        if num.method == ROUND4_METHOD:
                            kind = 'float'
                        else:
                            kind = 'count'
                            if num.rounded == LESS_THAN_15:
                                extra_spaces = len(LESS_THAN_15) - len(num.original)
                                if helpers.all_spaces(line[col1:col1 + extra_spaces]):
                                    col1 = col1 + extra_spaces

                        fbefore.write(line[pos:col0])  # part of line before span in question
                        fbefore.write("<span class='b{}'>".format(kind))
                        fbefore.write(line[col0:col1])  #  span in question
                        fbefore.write("</span>")

                        fafter.write(line[pos:col0])  # part of line after span in question
                        fafter.write("<span class='r{}'>".format(kind))
                        fafter.write(num.rounded)  # span in question
                        fafter.write("</span>")

                        frounded.write(line[pos:col0])
                        frounded.write(num.rounded)

                        pos = col1 # next character on line to process is col1

                # get end of line (or entire line, if not spans)
                fbefore.write(line[pos:])
                fafter.write(line[pos:])
                frounded.write(line[pos:])

        for what in [fbefore,fafter]:
            what.write("</div></body></html>\n")

        frounded.close()
        fbefore.close()
        fafter.close()

    def process(self):
        if self.args.log:
            self.process_logfile()
            return

        (base, ext) = os.path.splitext(self.fname)
        ext = ext.lower()
        if ext == ".xlsx":
            self.process_xlsx(is_xlsx=True)
        elif ext in [".xls", ".ods"]:
            self.process_xlsx(is_xlsx=False)
        elif ext in [".docx", ".odt"]:
            helpers.convert_word_to_txt(self.fname)
            (base, ext) = os.path.splitext(fname)
            self.fname = base + ".txt"
            self.process_logfile()
        elif ext in ['.csv','.tsv']:
            self.process_csvfile()
        elif ext in self.LOGFILE_ROUNDER_EXTENSIONS:
            self.process_logfile()
        else:
            rounder_logger.error("ABORT: Don't know how to process '{}' type file in: {}".format(ext, fname))
            sys.exit(1)


if __name__=="__main__":
    import argparse
    import getpass

    parser = argparse.ArgumentParser(description='Perform DRB-style rounding',
                                     formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("files", type=str, nargs="+", help="File[s] to round")
    parser.add_argument("--tab",    action="store_true", help="CSV is tab delimited")
    parser.add_argument("--highlight", action="store_true", help="Do not round spreadsheet values, only highlight")
    parser.add_argument("--log", help="Invoke logfile rounder. This is the default for files ending in {}"
                        .format(" ".join(DRBRounder.LOGFILE_ROUNDER_EXTENSIONS)))
    parser.add_argument("--counts", action='store_true', help="Apply rounding rules for small counts: 0-7 rounds to 4")
    parser.add_argument("--zap",    action='store_true', help="Overwrite output files")

    args = parser.parse_args()
    if args.tab == True:
        args.delimiter = "\t"
    else:
        args.delimiter = ","  # default delimiter

    # rounder_log is the logfile for all uses of the rounder. 
    rounder_logger = setup_logger('rounder', os.getcwd() + '/rounder.log', '%(asctime)s:\t%(message)s', 'a', True)

    for fname in args.files:
        rounder_logger.info("RUN: {} ran the rounder on '{}'" .format(getpass.getuser(), fname))

        d = DRBRounder(args, fname)
        d.process()
        rounder_logger.info("STATUS:   Logfile written to {}".format(d.logfile_path))
        rounder_logger.info("COMPLETE: {} ran the rounder on '{}'" .format(getpass.getuser(), fname))
