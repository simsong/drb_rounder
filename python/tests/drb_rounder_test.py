import pytest
import sys
import os
import os.path
import shutil

import subprocess
#
# Common tests for the DRB_ROUNDER
#

from common import *

sys.path.append( os.path.join(os.path.dirname(__file__),".."))

from number import Number


DRB_ROUNDER_PY= os.path.join(os.path.dirname(__file__),"../drb_rounder.py")
INFILES=[ ("concentration_stats.xlsx","concentration_stats_rounded.xlsx") ] 



DRB_ROUNDER_PY = os.path.join( os.path.dirname(__file__), "../drb_rounder.py")

def test_process_csv():
    """
    Test the process_csv function by doing the following:
        - Create a working directory for all test files
        - Move csv files into working directory
        - Verify contents of the csv file are what we expect
            - Read line by line and add to a list
        - Run the rounder
        - Verify that the contents of the outputted csv files are what we expect
    """
    import argparse

    FNAME_COMMA         = os.path.join(WORK_DIR, "test_comma.csv")
    FNAME_TAB           = os.path.join(WORK_DIR, "test_tab.csv")
    FNAME_COMMA_ROUNDED = os.path.join(WORK_DIR, "test_comma_rounded.csv")
    FNAME_TAB_ROUNDED   = os.path.join(WORK_DIR, "test_tab_rounded.csv")
    COMMA_ORIGINALS = ["41", "74.0", "170.1111111", "142", "689", "16612.832", "32.6", "7003200", "15", "10032", "0.5", "167"]
    COMMA_ROUNDED   = ["40", "74.0", "170.1",       "150", "700", "16610.",    "32.6", "7003000", "20", "10000", "0.5", "150"]

    TAB_ORIGINALS   = ["10",  "22", "6700.32", "500932", "1007382", "55.2"]
    TAB_ROUNDED     = ["<15", "20", "6700.",   "501000", "1007000", "55.2"]

    # First, make sure that each of these get rounded as expected.
    for i in range(len(COMMA_ORIGINALS)):
        n = Number( COMMA_ORIGINALS[i] )
        assert n.round().strip() == COMMA_ROUNDED[i].strip()

    for i in range(len(TAB_ORIGINALS)):
        n = Number( TAB_ORIGINALS[i] )
        assert n.round().strip() == TAB_ROUNDED[i].strip()


    prep_workdir()

    def read_csv(fname, delimiter):
        """
        Place all the digits inside a csv file into a list
        """
        result = []
        print("reading: ",fname)
        with open( fname,  "rU") as infile:
            for line in infile:
                print(line.strip())
                fields = line.split(delimiter)
                for field in fields:
                    # strip everything but digits and periods and <
                    new_field = ''.join(i for i in field if i.isdigit() or i=='.' or i=='<')  
                    if new_field != '':
                        result.append(new_field)
        return result

    # Make sure the original copies of .csv files are what we expect
    assert read_csv(FNAME_COMMA, ",") == COMMA_ORIGINALS
    assert read_csv(FNAME_TAB, "\t") == TAB_ORIGINALS
    
    # Run the rounder on both files

    assert os.path.exists(DRB_ROUNDER_PY)
    r = subprocess.call([sys.executable,DRB_ROUNDER_PY,'--zap',FNAME_COMMA])
    assert r==0
    assert read_csv(FNAME_COMMA_ROUNDED, ",") == COMMA_ROUNDED


    assert os.path.exists(DRB_ROUNDER_PY)
    r = subprocess.call([sys.executable,DRB_ROUNDER_PY,'--zap','--tab',FNAME_TAB])
    assert r==0
    assert read_csv(FNAME_TAB_ROUNDED,  "\t") == TAB_ROUNDED


def test_process_xlsx():
    """
    Tests the process_xlsx function by doing the following:
        - Create a working directory for all test files
        - Move test spreadsheets into this working directory
        - Verify that the contents of these spreadsheets are what we expect
        - Run the spreadsheet through the DRB Rounder
        - Verify that the contents of the outputted spreadsheets are also what
          we expect
    """
    import shutil
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.comments import Comment

    EXCEL_FN = os.path.join(WORK_DIR, "test.xlsx")
    ROUNDED_EXCEL_FN = os.path.join(WORK_DIR, "test_rounded.xlsx")

    FILL_NONE = PatternFill(start_color='00000000', end_color='00000000')
    FILL_ORANGE = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    FILL_BLUE = PatternFill(start_color='00b8ff', end_color='00b8ff', fill_type='solid')
    COMMENT_FLOAT = Comment('Orange: rounder identified a FLOAT that needed to be rounded', 'DRB_ROUNDER')
    COMMENT_INTEGER = Comment('Blue: rounder identified a COUNT that needed to be rounded', 'DRB_ROUNDER')

    prep_workdir()

    # Load Workbook
    assert os.path.exists(EXCEL_FN)
    wb = load_workbook(EXCEL_FN, data_only=True)
    ws = wb.active

    # Test to make sure the excel file is starting at the right state
    assert ws['A1'].value == "Heading 1"
    assert ws['B1'].value == "Heading 2"
    assert ws['C1'].value == "Heading 3"
    assert ws['A2'].value == 1
    assert ws['B2'].value == 1.01
    assert ws['C2'].value == 1.111111
    assert ws['A3'].value == 17
    assert ws['B3'].value == 17.01
    assert ws['C3'].value == 17.111111
    assert ws['A4'].value == 150
    assert ws['B4'].value == 150.01
    assert ws['C4'].value == 150.111111
    assert ws['A5'].value == 1250
    assert ws['B5'].value == 1250.01
    assert ws['C5'].value == 1250.111111

    for cell in ws.get_cell_collection():
        assert cell.fill == FILL_NONE

    wb.save(EXCEL_FN)

    # Run the rounder
    r = subprocess.call([sys.executable, DRB_ROUNDER_PY, '--zap', EXCEL_FN])
    assert r==0

    # Load Workbook
    wb = load_workbook(ROUNDED_EXCEL_FN, data_only=True)
    ws = wb.active

    # Check the end state of the rounder
    assert ws['A1'].value == "Heading 1"
    assert ws['B1'].value == "Heading 2"
    assert ws['C1'].value == "Heading 3"
    assert ws['A2'].value == "<15"
    assert ws['B2'].value == 1.01
    assert ws['C2'].value == 1.111
    assert ws['A3'].value == 20
    assert ws['B3'].value == 17.01
    assert ws['C3'].value == 17.11
    assert ws['A4'].value == 150
    assert ws['B4'].value == 150
    assert ws['C4'].value == 150.1
    assert ws['A5'].value == 1300
    assert ws['B5'].value == 1250
    assert ws['C5'].value == 1250

    assert ws['A1'].fill == FILL_NONE
    assert ws['B1'].fill == FILL_NONE
    assert ws['C1'].fill == FILL_NONE
    assert ws['A2'].fill == FILL_BLUE
    assert ws['B2'].fill == FILL_NONE
    assert ws['C2'].fill == FILL_ORANGE
    assert ws['A3'].fill == FILL_BLUE
    assert ws['B3'].fill == FILL_NONE
    assert ws['C3'].fill == FILL_ORANGE
    assert ws['A4'].fill == FILL_NONE
    assert ws['B4'].fill == FILL_ORANGE
    assert ws['C4'].fill == FILL_ORANGE
    assert ws['A5'].fill == FILL_BLUE
    assert ws['B5'].fill == FILL_ORANGE
    assert ws['C5'].fill == FILL_ORANGE

    assert ws['A2'].comment == COMMENT_INTEGER
    assert ws['C2'].comment == COMMENT_FLOAT



def test_files():
    """For specific files, run the rounder on them and verify the results."""
    for (infile,outfile) in INFILES:
        infile_path = os.path.join(WORK_DIR, infile)
        assert os.path.exists(infile_path)

        assert os.path.exists(DRB_ROUNDER_PY)
        res = subprocess.call([sys.executable,DRB_ROUNDER_PY, '--zap',infile_path])

