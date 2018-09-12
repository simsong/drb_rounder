import pytest
import sys
import pytest
import sys
import os
import os.path

from common import *
from helpers import *  # pylint: disable=W0614


################################################################
### String manipulation code
################################################################
def test_number_re():
    assert number_re.search("1") != None
    assert number_re.search("12") != None
    assert number_re.search(" 1 ") != None
    assert number_re.search(" 12 ") != None
    assert number_re.search(" 123 ") != None
    assert number_re.search(" 1234 ") != None
    assert number_re.search(" 1,234 ") != None
    assert number_re.search(" 11,234 ") != None
    assert number_re.search(" 123,234 ") != None
    assert number_re.search(" 1,123,234 ") != None


def test_find_next_number():
    assert find_next_number("1") == (0,1)
    assert find_next_number("1.") == (0,2)
    assert find_next_number("1.1") == (0,3)
    assert find_next_number(".123") == (0,4)
    assert find_next_number("abc.123") == (3,7)
    assert find_next_number("200") == (0,3)
    assert find_next_number("1,234") == (0,5)
    assert find_next_number("1234") == (0,4)
    assert find_next_number("123456") == (0,6)
    assert find_next_number("123,456") == (0,7)
    assert find_next_number("1234567") == (0,7)
    assert find_next_number("1234567,345") == (0,7)
    assert find_next_number(",") == None
    assert find_next_number(" 1, ") == (1,2)


def test_extract_number():
    assert extract_number("3.0")=="3.0"
    assert extract_number("[3.0]")=="3.0"
    assert extract_number(" [3.0] ")=="3.0"
    assert extract_number(" [3.0] ")=="3.0"
    assert extract_number("3.0e10 ")=="3.0e10"
    assert extract_number("3.0e-10 ")=="3.0e-10"

def test_numbers_in_line():
    line = " 1 2 3,456 7.890"
    positions = [(1,2), (3,4), (5,10), (11,16)]
    assert numbers_in_line(line) == positions


def test_in_scientific_form():
    assert in_scientific_form("1.234e-4") == True
    assert in_scientific_form("1.234e+4") == True
    assert in_scientific_form("+1.234e+4") == True
    assert in_scientific_form("-1.234e+4") == True
    assert in_scientific_form("+1.234e4") == True
    assert in_scientific_form("-1.234e4") == True
    assert in_scientific_form(".234E+4") == True
    assert in_scientific_form(".2E+40") == True
    assert in_scientific_form("2E+3") == True
    assert in_scientific_form("E+3") == False
    assert in_scientific_form("2E") == False


def test_num_trailing_zeros():
    assert num_trailing_zeros("1234.") == 0
    assert num_trailing_zeros("1234.0") == 1
    assert num_trailing_zeros("1234.00") == 2
    assert num_trailing_zeros("1234.000") == 3
    assert num_trailing_zeros("1234.0000") == 4
    assert num_trailing_zeros("1234.00000") == 5
    assert num_trailing_zeros(".123000") == 3
    assert num_trailing_zeros(".123001") == 0
    assert num_trailing_zeros(".123010") == 1


def test_find_sigfigs():
    assert find_sigfigs("1") == 1
    assert find_sigfigs("12") == 2
    assert find_sigfigs("123") == 3
    assert find_sigfigs(".1") == 1
    assert find_sigfigs("1.2") == 2
    assert find_sigfigs("1.23") == 3
    assert find_sigfigs("1.230") == 3
    assert find_sigfigs("1.2300") == 3
    assert find_sigfigs("1.23000") == 3
    assert find_sigfigs("1.230000") == 3


def test_remove_trailing_zero():
    assert remove_trailing_zero("1.0") == "1."
    assert remove_trailing_zero("12.0") == "12."
    assert remove_trailing_zero("123.0") == "123."
    assert remove_trailing_zero("123.00") == "123.00"



################################################################
### File Manipulation Code
################################################################
def test_convert_to_xlsx():
    """
    Test the convert_to_xlsx function by checking whether the initial xls
    sheet has the same values as the xlsx sheet.
    """
    if sys.platform != 'win32':
        return

    import shutil
    import xlrd
    from openpyxl import load_workbook

    # Remove directory if it exists and create new working directory
    prep_workdir()
    
    EXCEL_FN  = "test_xls.xls"
    EXCELX_FN = "test_xls.xlsx"

    # Copy excel files from test directory over to our work directory
    shutil.copy( os.path.join(TEST_FILES_DIR, EXCEL_FN),
                 os.path.join(WORK_DIR, EXCEL_FN))

    # Open xls workbook and check values in file
    xlsBook = xlrd.open_workbook( os.path.join(WORK_DIR, EXCEL_FN))
    xlsSheet = xlsBook.sheet_by_index(0)
    assert xlsSheet.cell_value(0,0) == 1
    assert xlsSheet.cell_value(1,0) == 2.4893
    assert xlsSheet.cell_value(2,0) == "head"
    assert xlsSheet.cell_value(3,0) == 4000000000
    assert xlsSheet.cell_value(4,0) == "5*"
    assert xlsSheet.cell_value(5,0) == "39()"
    assert xlsSheet.cell_value(0,1) == 7.2
    assert xlsSheet.cell_value(1,1) == 64
    assert xlsSheet.cell_value(2,1) == "tail"
    assert xlsSheet.cell_value(3,1) == 10897932

    convert_to_xlsx( os.path.join(WORK_DIR, EXCEL_FN))

    # Open xlsx workbook and check values in file
    wb = load_workbook( os.path.join(WORK_DIR, EXCELX_FN), data_only=True)
    ws = wb.active
    assert ws['A1'].value == 1
    assert ws['A2'].value == 2.4893
    assert ws['A3'].value == "head"
    assert ws['A4'].value == 4000000000
    assert ws['A5'].value == "5*"
    assert ws['A6'].value == "39()"
    assert ws['B1'].value == 7.2
    assert ws['B2'].value == 64
    assert ws['B3'].value == "tail"
    assert ws['B4'].value == 10897932


def test_check_for_excel_formulas():
    """
    Tests the check_for_excel_formulas by checking whether each sheet in the
    excel file has or doesn't have a formula
    """
    import shutil
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    TEST_FN              = "test_formula.xlsx"
    NOT_ROUNDED_TEST_FN = "test_formula_notrounded.xlsx"
    FILL_RED = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Remove directory if it exists and create new working directory
    prep_workdir()

    # Copy excel file from test directory over to our work directory
    shutil.copy( os.path.join(TEST_FILES_DIR, TEST_FN), WORK_DIR)

    # Make sure we see that there is a formula here
    assert check_for_excel_formulas( os.path.join(WORK_DIR, TEST_FN)) == True

    # Make sure each sheet has formulas where expected
    wb = load_workbook( os.path.join(WORK_DIR, NOT_ROUNDED_TEST_FN))
    ws = wb.get_sheet_by_name("Sheet2")
    assert str(ws["H12"].value)[0] == "="  # this symbol means it is a formula
    assert ws["H12"].fill == FILL_RED

    ws = wb.get_sheet_by_name("Sheet3")
    assert str(ws["G4"].value)[0] == "="
    assert ws["G4"].fill == FILL_RED

    ws = wb.get_sheet_by_name("Sheet4")
    assert str(ws["G5"].value)[0] == "="
    assert ws["G5"].fill == FILL_RED


################################################################
### Rounding code
################################################################
def test_nearest():
    assert nearest(80,100) == 100
    assert nearest(120,100) == 100
    assert nearest(51,100) == 100
    assert nearest(51,10) == 50
    assert nearest(56,10) == 60
    assert nearest(60,10) == 60
    assert nearest(61,10) == 60


def test_round_counts():
    assert round_counts(1) == "<15"
    assert round_counts(14) == "<15"
    assert round_counts(15) == "20"
    assert round_counts(20) == "20"
    assert round_counts(24) == "20"


def test_round4_float():
    assert round4_float(1.0)    == 1.0
    assert round4_float(1.2)    == 1.2
    assert round4_float(1.23)   == 1.23
    assert round4_float(1.234)  == 1.234
    assert round4_float(1.2345) == 1.234
    assert round4_float(10)    == 10
    assert round4_float(12)    == 12
    assert round4_float(12.3)   == 12.3
    assert round4_float(12.34)  == 12.34
    assert round4_float(12.345) == 12.35


def test_round4_decimal():
    assert round4_decimal(1234)  == 1234
    assert round4_decimal(12345) == 12340
    assert round4_decimal(12346) == 12350
    assert round4_decimal(123450) == 123400
    assert round4_decimal(123455) == 123500
    assert round4_decimal(123465) == 123500

