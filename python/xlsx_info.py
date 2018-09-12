#!/usr/bin/env python3

"""
Prepare a report of disclosure avoidance statistics about an excel spreadsheet.
Output can be in HTML, TEXT or LATEX.
Can be run stand-alone or within the combined file printer.
"""

from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell
from collections import defaultdict
import zipfile
import tytable
from helpers import extract_number
from tytable import ttable, HTML, TEXT, LATEX

from latex_tools import latex_escape, textbf, nl

import os
import os.path
import re
import sys
sigfig_re = re.compile("([0-9.]+)")

ALERT_DIGITS = 5                # alert if this many digits or more
ALERT_VALUES = 10               # alert on this many values

def colname(col):
    assert 0<=col<=26+26*26     # only handles 1 and 2 character cols

    col -= 1                    # column 0 is A
    if col<26:
        return chr( ord("A") + col)
    r1 = (col-26) // 26
    r2 = (col-26) % 26
    return colname(r1) + colname(r2)

def improperly_rounded(val):
    return SigFigStats.find_sigfigs(val) >= ALERT_DIGITS


class ProblemStat:
    def __init__(self,*,sheet="",row="",col="",source="xls"):
        self.sheet = sheet
        self.row   = row
        self.col   = col
        self.source  = source

    def __lt__(self,v):
        if self.row < v.row: return True
        if self.row > v.row: return False
        if self.col < v.col: return True
        return False
        
    def __le__(self,v):
        if self.row <= v.row: return True
        if self.row > v.row: return False
        if self.col <= v.col: return True
        return False
        
    def __str__(self):
        if self.source=="xls":
            if self.sheet:
                return "{}:{}{}".format(self.sheet,colname(self.col),self.row)
            return "{}{}".format(colname(self.col),self.row)

class SigFigStats:
    """Maintain statistics about improperly rounded values"""
    @staticmethod
    def find_sigfigs(x):
        """Odd little function to return the number of significant figures in a number"""
        # Kill everything to the right of the e and remove the period
        m = sigfig_re.search("{:e}".format(float(x)))
        if m:
            val = m.group(1).replace(".","")
            while val[-1:]=='0':
                val = val[0:-1]
            return len(val)
        return 0                    # no significant figures if it can't be made scientific

    def __init__(self):
        self.max_row = 0
        self.max_col = 0
        self.sig_digits_histogram = defaultdict(int) # count
        self.sig_digits_alerts    = defaultdict(list)
        self.count = 0
        self.improper_count = 0

    def add(self,*, val, row, col):
        """Adds a number to the counter and return True if it has more than ALERT_DIGITS significant figures"""
        sigfigs = SigFigStats.find_sigfigs(val)
        self.max_row = max(self.max_row, row)
        self.max_col = max(self.max_col, col)
        self.sig_digits_histogram[sigfigs] += 1
        self.count += 1
        if improperly_rounded(val):
            self.improper_count += 1
            self.sig_digits_alerts[sigfigs].append(ProblemStat(row=row,col=col))

    def typeset(self,*,mode):
        if not self.count:
            return ""
        t = ttable()
        t.add_head(['Significant Digits','Cell count','Problem cells'])
        t.set_latex_colspec("rrl")
        t.set_col_alignment(1,ttable.CENTER)
        for (sig_digs,vals) in sorted(self.sig_digits_alerts.items()):
            alert_cells = sorted(vals)[0:ALERT_VALUES]
            alerts = " ".join( [str(a) for a in alert_cells])
            if len(alert_cells) >= ALERT_VALUES:
                alerts += r" . . . "
            t.add_data( (sig_digs,len(vals),alerts) )
        if t.data:
            return t.typeset(mode=mode)
        else:
            return f"Cells: {self.count}; nothing improperly rounded."
    
def get_number(s):
    try:
        return float(s)
    except (ValueError,TypeError):
        pass
    try:
        return float(extract_number(s))
    except (ValueError,TypeError):
        pass
    return None

# A tool for reading CSV files that matches the openpyxl interface
class CSVCell:
    def __init__(self,rownumber,column,value):
        self.row    = rownumber
        self.column = column
        self.value   = value
    def __str__(self):
        return f"<CSVCell {self.row},{self.column},{self.value}>"

import csv
class CSVFile:
    def __init__(self,fname):
        import csv
        self.fname = fname
        self.title = os.path.basename(fname)

    def iter_rows(self):
        """Return an iterator that is an array of cells"""
        with open(self.fname,"r") as csvfile:
            linenumber = 0
            for row in csv.reader(csvfile, delimiter=','):
                linenumber += 1
                ret = []
                for column in range(len(row)):
                    ret.append( CSVCell(rownumber=linenumber,column=column+1,value=row[column]) )
                yield ret
                

class CSVWorkbook:
    def __init__(self,filename):
        self.worksheets = [CSVFile(filename)]
        self.active     = self.worksheets[0]
    
def load_csv_workbook(*,filename,read_only=True):
    assert read_only==True
    return CSVWorkbook(filename)
        


def analyze_file(*,filename,mode=TEXT):
    """Analyze the named excel spreadsheet and return a string in the specified format."""
    NL = tytable.line_end(mode)
    ret = [NL]                   # an array of strings to be joined and returned
    if mode==TEXT:
        ret += [filename,'\n',"="*len(filename),NL]
    elif mode==LATEX:
        ret += [textbf(latex_escape(filename+":")),NL]
        
    if filename.lower().endswith(".csv"):
        wb = load_csv_workbook(filename=filename)
    elif filename.lower().endswith(".xlsx"):
        try:
            wb = load_workbook(filename=filename, read_only=True)
        except zipfile.BadZipFile:
            ret += ['Cannot read; badZipFile',NL]
            return "".join(ret)
        except TypeError:
            ret += ['Cannot read; bad Excel File',NL]
            return "".join(ret)
    else:
        return "Cannot read file type: {}".format(filename)

    tt = ttable()
    tt.add_head(['Worksheet Name','rows with data','columns with data',
                 'total numeric values','cells with $>4$ sigfigs'])
    
    ret_worksheets = []
    for ws in wb.worksheets:
        print(f"  Analyzing {os.path.basename(filename)} worksheet {ws.title}")
        sf_ws = SigFigStats()              # for this worksheet
        
        empty_cells = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell,EmptyCell):
                    empty_cells += 1
                    continue
                val = get_number(cell.value)
                if val!=None:
                    sf_ws.add(val=val, row=cell.row, col=cell.column)
        ret_worksheets += ["Worksheet "+latex_escape(ws.title)]
        ret_worksheets += [sf_ws.typeset(mode=mode)]
        ret_worksheets += ["rows: {}  columns: {}  numbers: {}".format(sf_ws.max_row,sf_ws.max_col,sf_ws.count)]
        ret_worksheets += [""]
        tt.add_data([latex_escape(ws.title),sf_ws.max_row,sf_ws.max_col,sf_ws.count, sf_ws.improper_count])
        # End of worksheet processing

    tt.set_col_totals([3,4])
    ret += [tt.typeset(mode=mode)] + [NL] + ret_worksheets

    # count the number of numbers and their precision
    return NL.join(ret)

def can_analyze_file(*,filename):
    return filename.lower().endswith(".xlsx") or filename.lower().endswith(".csv")

def mytest_annotate(fname):
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Color, PatternFill, Font, Border

    wb = load_workbook(fname)
    ws = wb.active
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    ws['A1'].comment = Comment("Comment for A1","No Such Author")
    ws['A2'].fill = redFill
    (base,ext) = os.path.splitext(fname)
    wb.save(base+"-new"+ext)
 

if __name__=="__main__":
    from argparse import ArgumentParser,ArgumentDefaultsHelpFormatter
    parser = ArgumentParser( formatter_class = ArgumentDefaultsHelpFormatter )
    parser.add_argument("file",help="file to analyze")
    parser.add_argument("--test", help="Test a file by making minor modificaitons to it.", action='store_true')
    parser.add_argument("--debug", action='store_true')
    parser.add_argument("--text", action='store_true')
    parser.add_argument("--latex", action='store_true')
    args = parser.parse_args()

    if args.test:
        mytest_annotate(args.file)
        exit(1)

    if not args.text and not args.latex:
        args.text = True

    if args.text:
        print(analyze_file(filename=args.file,mode=TEXT))
        print("")

    if args.latex:
        print(analyze_file(filename=args.file,mode=LATEX))
        print("")
